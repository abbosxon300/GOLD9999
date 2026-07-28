from datetime import date, timedelta
from io import BytesIO

from flask import render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def register_reports_routes(
    app,
    *,
    app_name,
    init_db,
    get_db,
    fmt_uzs,
    login_required,
    admin_required,
):
    @app.route("/reports")
    @login_required
    @admin_required
    def reports():
        return render_template(
            "reports.html",
            app_name=app_name,
        )

    @app.route("/reports/profit")
    @login_required
    @admin_required
    def report_profit():
        init_db()
        db = get_db()

        today = date.today()
        default_from = (today - timedelta(days=30)).isoformat()
        default_to = today.isoformat()

        from_date = (
            request.args.get("from") or default_from
        ).strip()

        to_date = (
            request.args.get("to") or default_to
        ).strip()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        summary = db.execute("""
            SELECT
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
        """, (
            from_date,
            to_date,
        )).fetchone()

        rows = db.execute("""
            SELECT
                sale_date,
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
            GROUP BY sale_date
            ORDER BY sale_date DESC
        """, (
            from_date,
            to_date,
        )).fetchall()

        return render_template(
            "report_profit.html",
            app_name=app_name,
            from_date=from_date,
            to_date=to_date,
            sales_count=int(summary["sales_count"] or 0),
            total_sell_fmt=fmt_uzs(summary["total_sell"] or 0),
            total_cost_fmt=fmt_uzs(summary["total_cost"] or 0),
            total_profit_fmt=fmt_uzs(summary["total_profit"] or 0),
            rows=rows,
        )


    @app.route("/reports/agents")
    @login_required
    @admin_required
    def report_agents():
        init_db()
        db = get_db()

        today = date.today()
        default_from = (today - timedelta(days=30)).isoformat()
        default_to = today.isoformat()

        from_date = (
            request.args.get("from") or default_from
        ).strip()

        to_date = (
            request.args.get("to") or default_to
        ).strip()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        rows = db.execute("""
            SELECT
                s.agent_id,
                COALESCE(
                    NULLIF(TRIM(u.full_name), ''),
                    u.username,
                    'Noma’lum'
                ) AS agent_name,
                COUNT(*) AS sales_count,
                COALESCE(SUM(s.total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(s.total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(s.total_profit_uzs), 0) AS total_profit
            FROM sales s
            LEFT JOIN users u
              ON u.id=s.agent_id
            WHERE s.sale_date BETWEEN ? AND ?
            GROUP BY
                s.agent_id,
                agent_name
            ORDER BY
                total_profit DESC,
                total_sell DESC,
                agent_name
        """, (
            from_date,
            to_date,
        )).fetchall()

        totals = {
            "sales_count": sum(
                int(row["sales_count"] or 0)
                for row in rows
            ),
            "total_sell": sum(
                float(row["total_sell"] or 0)
                for row in rows
            ),
            "total_cost": sum(
                float(row["total_cost"] or 0)
                for row in rows
            ),
            "total_profit": sum(
                float(row["total_profit"] or 0)
                for row in rows
            ),
        }

        return render_template(
            "report_agents.html",
            app_name=app_name,
            from_date=from_date,
            to_date=to_date,
            rows=rows,
            sales_count=totals["sales_count"],
            total_sell_fmt=fmt_uzs(totals["total_sell"]),
            total_cost_fmt=fmt_uzs(totals["total_cost"]),
            total_profit_fmt=fmt_uzs(totals["total_profit"]),
        )


    @app.route("/reports/daily")
    @login_required
    @admin_required
    def report_daily():
        init_db()
        db = get_db()

        today = date.today()
        default_from = (today - timedelta(days=30)).isoformat()
        default_to = today.isoformat()

        from_date = (
            request.args.get("from") or default_from
        ).strip()

        to_date = (
            request.args.get("to") or default_to
        ).strip()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        rows = db.execute("""
            SELECT
                sale_date,
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
            GROUP BY sale_date
            ORDER BY sale_date DESC
        """, (
            from_date,
            to_date,
        )).fetchall()

        totals = {
            "sales_count": sum(
                int(row["sales_count"] or 0)
                for row in rows
            ),
            "total_sell": sum(
                float(row["total_sell"] or 0)
                for row in rows
            ),
            "total_cost": sum(
                float(row["total_cost"] or 0)
                for row in rows
            ),
            "total_profit": sum(
                float(row["total_profit"] or 0)
                for row in rows
            ),
        }

        return render_template(
            "report_daily.html",
            app_name=app_name,
            from_date=from_date,
            to_date=to_date,
            rows=rows,
            sales_count=totals["sales_count"],
            total_sell_fmt=fmt_uzs(totals["total_sell"]),
            total_cost_fmt=fmt_uzs(totals["total_cost"]),
            total_profit_fmt=fmt_uzs(totals["total_profit"]),
        )


    @app.route("/reports/monthly")
    @login_required
    @admin_required
    def report_monthly():
        init_db()
        db = get_db()

        today = date.today()
        default_from = today.replace(month=1, day=1).isoformat()
        default_to = today.isoformat()

        from_date = (
            request.args.get("from") or default_from
        ).strip()

        to_date = (
            request.args.get("to") or default_to
        ).strip()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        rows = db.execute("""
            SELECT
                SUBSTR(sale_date, 1, 7) AS sale_month,
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
            GROUP BY SUBSTR(sale_date, 1, 7)
            ORDER BY sale_month DESC
        """, (
            from_date,
            to_date,
        )).fetchall()

        totals = {
            "sales_count": sum(
                int(row["sales_count"] or 0)
                for row in rows
            ),
            "total_sell": sum(
                float(row["total_sell"] or 0)
                for row in rows
            ),
            "total_cost": sum(
                float(row["total_cost"] or 0)
                for row in rows
            ),
            "total_profit": sum(
                float(row["total_profit"] or 0)
                for row in rows
            ),
        }

        return render_template(
            "report_monthly.html",
            app_name=app_name,
            from_date=from_date,
            to_date=to_date,
            rows=rows,
            sales_count=totals["sales_count"],
            total_sell_fmt=fmt_uzs(totals["total_sell"]),
            total_cost_fmt=fmt_uzs(totals["total_cost"]),
            total_profit_fmt=fmt_uzs(totals["total_profit"]),
        )


    @app.route("/reports/excel")
    @login_required
    @admin_required
    def reports_excel():
        init_db()
        db = get_db()

        today = date.today()
        default_from = today.replace(month=1, day=1).isoformat()
        default_to = today.isoformat()

        from_date = (
            request.args.get("from") or default_from
        ).strip()

        to_date = (
            request.args.get("to") or default_to
        ).strip()

        if from_date > to_date:
            from_date, to_date = to_date, from_date

        summary = db.execute("""
            SELECT
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
        """, (
            from_date,
            to_date,
        )).fetchone()

        daily_rows = db.execute("""
            SELECT
                sale_date,
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
            GROUP BY sale_date
            ORDER BY sale_date
        """, (
            from_date,
            to_date,
        )).fetchall()

        monthly_rows = db.execute("""
            SELECT
                SUBSTR(sale_date, 1, 7) AS sale_month,
                COUNT(*) AS sales_count,
                COALESCE(SUM(total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(total_profit_uzs), 0) AS total_profit
            FROM sales
            WHERE sale_date BETWEEN ? AND ?
            GROUP BY SUBSTR(sale_date, 1, 7)
            ORDER BY sale_month
        """, (
            from_date,
            to_date,
        )).fetchall()

        agent_rows = db.execute("""
            SELECT
                COALESCE(
                    NULLIF(TRIM(u.full_name), ''),
                    u.username,
                    'Noma’lum'
                ) AS agent_name,
                COUNT(*) AS sales_count,
                COALESCE(SUM(s.total_sell_uzs), 0) AS total_sell,
                COALESCE(SUM(s.total_cost_uzs), 0) AS total_cost,
                COALESCE(SUM(s.total_profit_uzs), 0) AS total_profit
            FROM sales s
            LEFT JOIN users u
              ON u.id=s.agent_id
            WHERE s.sale_date BETWEEN ? AND ?
            GROUP BY
                s.agent_id,
                agent_name
            ORDER BY
                total_profit DESC,
                total_sell DESC,
                agent_name
        """, (
            from_date,
            to_date,
        )).fetchall()

        workbook = Workbook()
        workbook.remove(workbook.active)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        title_fill = PatternFill(
            fill_type="solid",
            fgColor="B4C6E7",
        )

        total_fill = PatternFill(
            fill_type="solid",
            fgColor="E2F0D9",
        )

        thin_side = Side(
            style="thin",
            color="B7B7B7",
        )

        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        money_format = '#,##0'

        def style_title(ws, title, columns):
            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=columns,
            )

            cell = ws.cell(1, 1, title)
            cell.font = Font(
                bold=True,
                size=14,
            )
            cell.fill = title_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            ws.row_dimensions[1].height = 26

        def style_headers(ws, row_number, columns):
            for column in range(1, columns + 1):
                cell = ws.cell(row_number, column)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        def style_table(ws, start_row, end_row, columns):
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=1,
                max_col=columns,
            ):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical="center",
                    )

        # ===== UMUMIY =====

        ws = workbook.create_sheet("Umumiy")

        style_title(
            ws,
            "GOLD 9999 — UMUMIY HISOBOT",
            4,
        )

        ws.append([
            "Davr",
            "Sotuvlar soni",
            "Jami sotuv",
            "Jami tannarx",
        ])

        ws.append([
            f"{from_date} — {to_date}",
            int(summary["sales_count"] or 0),
            float(summary["total_sell"] or 0),
            float(summary["total_cost"] or 0),
        ])

        ws.append([
            "",
            "",
            "Jami foyda",
            float(summary["total_profit"] or 0),
        ])

        style_headers(ws, 2, 4)
        style_table(ws, 3, 4, 4)

        ws["C4"].font = Font(bold=True)
        ws["D4"].font = Font(bold=True)
        ws["C4"].fill = total_fill
        ws["D4"].fill = total_fill

        for cell in ("C3", "D3", "D4"):
            ws[cell].number_format = money_format

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        ws.freeze_panes = "A2"

        # ===== KUNLIK =====

        ws = workbook.create_sheet("Kunlik")

        style_title(
            ws,
            "KUNLIK HISOBOT",
            5,
        )

        ws.append([
            "Sana",
            "Sotuvlar",
            "Sotuv",
            "Tannarx",
            "Foyda",
        ])

        for row in daily_rows:
            ws.append([
                row["sale_date"],
                int(row["sales_count"] or 0),
                float(row["total_sell"] or 0),
                float(row["total_cost"] or 0),
                float(row["total_profit"] or 0),
            ])

        style_headers(ws, 2, 5)

        if ws.max_row >= 3:
            style_table(ws, 3, ws.max_row, 5)

        for row_number in range(3, ws.max_row + 1):
            for column in ("C", "D", "E"):
                ws[f"{column}{row_number}"].number_format = money_format

        ws.auto_filter.ref = f"A2:E{max(ws.max_row, 2)}"
        ws.freeze_panes = "A3"

        widths = {
            "A": 16,
            "B": 14,
            "C": 18,
            "D": 18,
            "E": 18,
        }

        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        # ===== OYLIK =====

        ws = workbook.create_sheet("Oylik")

        style_title(
            ws,
            "OYLIK HISOBOT",
            5,
        )

        ws.append([
            "Oy",
            "Sotuvlar",
            "Sotuv",
            "Tannarx",
            "Foyda",
        ])

        for row in monthly_rows:
            ws.append([
                row["sale_month"],
                int(row["sales_count"] or 0),
                float(row["total_sell"] or 0),
                float(row["total_cost"] or 0),
                float(row["total_profit"] or 0),
            ])

        style_headers(ws, 2, 5)

        if ws.max_row >= 3:
            style_table(ws, 3, ws.max_row, 5)

        for row_number in range(3, ws.max_row + 1):
            for column in ("C", "D", "E"):
                ws[f"{column}{row_number}"].number_format = money_format

        ws.auto_filter.ref = f"A2:E{max(ws.max_row, 2)}"
        ws.freeze_panes = "A3"

        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        # ===== SOTUVCHILAR =====

        ws = workbook.create_sheet("Sotuvchilar")

        style_title(
            ws,
            "SOTUVCHILAR HISOBOTI",
            5,
        )

        ws.append([
            "Sotuvchi",
            "Sotuvlar",
            "Sotuv",
            "Tannarx",
            "Foyda",
        ])

        for row in agent_rows:
            ws.append([
                row["agent_name"],
                int(row["sales_count"] or 0),
                float(row["total_sell"] or 0),
                float(row["total_cost"] or 0),
                float(row["total_profit"] or 0),
            ])

        style_headers(ws, 2, 5)

        if ws.max_row >= 3:
            style_table(ws, 3, ws.max_row, 5)

        for row_number in range(3, ws.max_row + 1):
            for column in ("C", "D", "E"):
                ws[f"{column}{row_number}"].number_format = money_format

        ws.auto_filter.ref = f"A2:E{max(ws.max_row, 2)}"
        ws.freeze_panes = "A3"

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            f"gold9999_hisobot_"
            f"{from_date}_{to_date}.xlsx"
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )
